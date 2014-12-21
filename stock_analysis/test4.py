# -*- coding: gbk -*-
import urllib2

from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font

file_output = open(r'stock_number_list.txt','wt')

outwb = Workbook()

ws1 = outwb.create_sheet(0,u'sh_stock_data')

ws2 = outwb.create_sheet(0,u'sz_stock_data')

ft = Font(color=colors.RED)
s1 = Style(number_format='0%')
s2 = Style(number_format='0.00')

ws1.cell(row=1,column=1).value=u'股票编号'
ws1.cell(row=1,column=2).value=u'股票名称'
ws1.cell(row=1,column=3).value=u'涨跌幅'
ws1.cell(row=1,column=4).value=u'当前价格'

ws2.cell(row=1,column=1).value=u'股票编号'
ws2.cell(row=1,column=2).value=u'股票名称'
ws2.cell(row=1,column=3).value=u'涨跌幅'
ws2.cell(row=1,column=4).value=u'当前价格'

##ws1.write(0, 0, u'股票编号')
##ws1.write(0, 1, u'股票名称')
##ws1.write(0, 2, u'涨跌幅')
##ws1.write(0, 3, u'当前价格')
##
##ws2 = wb.add_sheet(u'sz_stock_data')
##ws2.write(0, 0, u'股票编号')
##ws2.write(0, 1, u'股票名称')
##ws2.write(0, 2, u'涨跌幅')
##ws2.write(0, 3, u'当前价格')

def get_one_stock_info(stock_list,row_count,sheet,useful_stock_num_list):
    #count = len(stock_list)
    stock_num = ','.join(stock_list)
    url = 'http://hq.sinajs.cn/list='+stock_num
    headers = {"User-Agent":"Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6"}
    req = urllib2.Request( url, headers = headers)
    content = urllib2.urlopen(req).read()
    feedback = content.decode('gbk')
    stocks = feedback.split(';')
    stocks = stocks[:-1]
    #useless_stock_num = []
    #useful_stock_num_list = []
    #for stock in stocks:
    for stock in stocks:
        #print stock
        #print "row_offset=%s,i=%s" % (row_offset,i)
        #print str(row_offset+i)
        data = stock.split('"')[1].split(',')
        if len(data)>1:
            useful_stock_num_list.append(process_per_stock(stock,row_count+i,sheet))
            row_count += 1
    return useful_stock_num_list,row_count



def process_per_stock(stock,row_count,sheet):
    data = stock.split('"')[1].split(',')
    name = '%-6s' % data[0]
    price_current = '%-6s' % float(data[3])
    change_percent = (float(data[3])-float(data[2]))*100/float(data[2])
    change_percent = '%-6s' % round(change_percent,2)
    #print("股票名称:{0} 涨跌幅:{1} 最新价:{2}".format(name,change_percent,price_current))
    #print "股票名称:%s 涨跌幅:%s 最新价:%s" % (name,change_percent,price_current
    #print str(row_number+1)
    end_point = stock.find('=')
    sheet.cell(row=row_count+1,column=1).value=stock[end_point-8:end_point]
    sheet.cell(row=row_count+1,column=2).value=name
    #sheet.cell(row=row_count+2,column=3).style.number_format.format_code = '0.00E+00' 
    sheet.cell(row=row_count+1,column=3).value=float(change_percent)
    sheet.cell(row=row_count+1,column=3).style = s1
    #sheet.cell(row=row_count+2,column=4).style.number_format.format_code = '0.00E+00' 
    sheet.cell(row=row_count+1,column=4).value=float(price_current)
    sheet.cell(row=row_count+1,column=3).style = s2
##        sheet.write(row_number+1, 0, stock[12:20])
##        sheet.write(row_number+1, 1, name)
##        sheet.write(row_number+1, 2, change_percent)
##        sheet.write(row_number+1, 3, price_current)
    print stock[end_point-8:end_point],name,change_percent,price_current
    return str(stock[end_point-8:end_point])
        
def generate_sotck_list(market_name,count_start,count_end):
    stock_list = []
    count = count_start
    while count < count_end:
        stock_list.append(market_name+str(count).zfill(6))
        count += 1
    return stock_list

each_count = 500
count_max = 602000

print "欢迎来到上海股市"
row_count = 1
sh_useful_stock_num_list = []
for i in range(4): #4
    #row_offset = i*each_count + 600000
    count_start = i*each_count + 600000
    count_end = (i+1)*each_count + 600000
    sh_stock = generate_sotck_list('sh',count_start,count_end)
    sh_useful_stock_num_list,row_count = get_one_stock_info(sh_stock,row_count,ws1,sh_useful_stock_num_list)

print "欢迎来到深圳股市"
row_count = 1
sz_useful_stock_num_list = []
for i in range(6): #6
    #row_offset = i*each_count
    count_start = i*each_count
    count_end = (i+1)*each_count
    sz_stock = generate_sotck_list('sz',count_start,count_end)
    sz_useless,row_count = get_one_stock_info(sz_stock,row_count,ws2,sz_useful_stock_num_list)
print "row_count"+str(row_count)
#row_offset = i*each_count
count_start = 300000
count_end = each_count + 300000
sz_stock = generate_sotck_list('sz',count_start,count_end)
sz_useful_stock_num_list,row_count = get_one_stock_info(sz_stock,row_count,ws2,sz_useful_stock_num_list)



#col_1_C = ws1.column_dimensions['C']
#col_1_C.style = s1
#col_1_D = ws1.column_dimensions['D']
#col_1_D.style = s2

#col_2_C = ws2.column_dimensions['C']
#col_2_C.style = s1
#col_2_D = ws2.column_dimensions['D']
#col_2_D.style = s2

new_filename = r'stock_basic3.xlsx'
outwb.save(new_filename)

print >>file_output,'sh_stock_num_list = ',sh_useful_stock_num_list,'\n\n\n','sz_stock_num_list = ',sz_useful_stock_num_list 
file_output.close()

##sh: 600000  602000
##sz: 000000  003000  300000  300500
