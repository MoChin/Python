# -*- coding: gbk -*-
import urllib2

from openpyxl import Workbook

outwb = Workbook()

ws1 = outwb.create_sheet(0,u'sh_stock_data')
ws2 = outwb.create_sheet(0,u'sz_stock_data')

ws1.cell(row=1,column=1).value=u'股票编号'
ws1.cell(row=1,column=2).value=u'股票名称'
ws1.cell(row=1,column=3).value=u'涨跌幅'
ws1.cell(row=1,column=4).value=u'当前价格'

ws2.cell(row=1,column=1).value=u'股票编号'
ws2.cell(row=1,column=2).value=u'股票名称'
ws2.cell(row=1,column=3).value=u'涨跌幅'
ws2.cell(row=1,column=4).value=u'当前价格'

##ws1.write(0, 0, u'股票编号')
##ws1.write(0, 1, u'股票名称')
##ws1.write(0, 2, u'涨跌幅')
##ws1.write(0, 3, u'当前价格')
##
##ws2 = wb.add_sheet(u'sz_stock_data')
##ws2.write(0, 0, u'股票编号')
##ws2.write(0, 1, u'股票名称')
##ws2.write(0, 2, u'涨跌幅')
##ws2.write(0, 3, u'当前价格')

def get_one_stock_info(stock_list,row_offset,sheet):
    #count = len(stock_list)
    stock_num = ','.join(stock_list)
    url = 'http://hq.sinajs.cn/list='+stock_num
    headers = {"User-Agent":"Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6"}
    req = urllib2.Request( url, headers = headers)
    content = urllib2.urlopen(req).read()
    feedback = content.decode('gbk')
    stocks = feedback.split(';')
    stocks = stocks[:-1]
    useless_stock_num = []
    #for stock in stocks:
    for i in range(len(stocks)):
        #print stock
        #print "row_offset=%s,i=%s" % (row_offset,i)
        #print str(row_offset+i)
        is_useful = process_per_stock(stocks[i],row_offset+i,sheet)
        if is_useful != None:
            useless_stock_num.append(str(is_useful)[-9:-3])
    return useless_stock_num



def process_per_stock(stock,row_number,sheet):
    data = stock.split('"')[1].split(',')
    if len(data)>1:
        name = '%-6s' % data[0]
        price_current = '%-6s' % float(data[3])
        change_percent = (float(data[3])-float(data[2]))*100/float(data[2])
        change_percent = '%-6s' % round(change_percent,2)
        #print("股票名称:{0} 涨跌幅:{1} 最新价:{2}".format(name,change_percent,price_current))
        #print "股票名称:%s 涨跌幅:%s 最新价:%s" % (name,change_percent,price_current
        #print str(row_number+1)
        sheet.cell(row=row_number+2,column=1).value=stock[12:20]
        sheet.cell(row=row_number+2,column=2).value=name
        sheet.cell(row=row_number+2,column=3).value=change_percent
        sheet.cell(row=row_number+2,column=4).value=price_current
##        sheet.write(row_number+1, 0, stock[12:20])
##        sheet.write(row_number+1, 1, name)
##        sheet.write(row_number+1, 2, change_percent)
##        sheet.write(row_number+1, 3, price_current)
        print name,change_percent,price_current
        #row_count += 1
    else:
        return stock
        
        

def generate_sotck_list(market_name,count_start,count_end):
    stock_list = []
    count = count_start
    while count < count_end:
        stock_list.append(market_name+str(count).zfill(6))
        count += 1
    return stock_list

each_count = 500
count_max = 620000

print "欢迎来到上海股市"
for i in range(4):
    row_offset = i*each_count + 600000
    count_start = i*each_count + 600000
    count_end = (i+1)*each_count + 600000
    sh_stock = generate_sotck_list('sh',count_start,count_end)
    sh_useless = get_one_stock_info(sh_stock,row_offset,ws1)

print "欢迎来到深圳股市"
for i in range(6):
    row_offset = i*each_count
    count_start = i*each_count
    count_end = (i+1)*each_count
    sz_stock = generate_sotck_list('sz',count_start,count_end)
    sz_useless = get_one_stock_info(sz_stock,row_offset,ws2)

new_filename = r'stock_basic.xlsx'
outwb.save(new_filename)


##sh: 600000  602000
##sz: 000000  003000  300000  300500
